#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
#   "docling>=2.0",
#   "python-dotenv>=1.0",
# ]
# ///
"""Build unified SQLite database from organizational data sources.

Loads:
- Org spreadsheet (XLSX) — normalized people, orgs, components, teams
- Jira issues export (CSV) — issues with labels/components
- Feature planning / RICE scores (CSV) — Jira Advanced Roadmaps export
- Release schedule (CSV) — release milestone dates
- Release-to-component version mapping (CSV) — unpivoted matrix
- Org chart (CSV) — raw hierarchical text

Usage:
    uv run scripts/build_db.py                # auto-find sources, build gps.db
    uv run scripts/build_db.py --force        # rebuild even if unchanged
"""

import argparse
import csv
import hashlib
import json
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
DATA_DIR = REPO_ROOT / "data"

try:
    from dotenv import load_dotenv

    load_dotenv(REPO_ROOT / ".env")
except ModuleNotFoundError:
    pass
VERSION_PATH = REPO_ROOT / "VERSION"
GOVERNANCE_DIR = REPO_ROOT / "governance"

# ---------------------------------------------------------------------------
# Org spreadsheet constants
# ---------------------------------------------------------------------------

DEFAULT_TAB_ORG_MAP = {
    "Platform - Team Breakdown": ("platform", "Platform"),
    "Data Services - Team Breakdown": ("data-services", "Data Services"),
    "ML Engineering - Team Breakdown": ("ml-engineering", "ML Engineering"),
}

TAB_ORG_MAP = json.loads(os.environ.get("GPS_TAB_ORG_MAP", json.dumps(DEFAULT_TAB_ORG_MAP)))

JIRA_SCRUM_REF_TAB = os.environ.get("GPS_JIRA_SCRUM_REF_TAB", "Jira and Scrum teams")

HEADER_ALIASES: dict[str, list[str]] = {
    "associate's name": ["associate's name"],
    "manager's name": ["manager's name"],
    "miro team name": ["miro team name", "scrum team name"],
    "primary jira component": ["primary jira component"],
    "jira team name": ["jira team name", "jira filter"],
    "pm": ["pm"],
    "eng lead": ["eng lead"],
    "status": ["status"],
    "engineering speciality": ["engineering speciality"],
    "last modified date": ["last modified date"],
}

EXPECTED_HEADERS = list(HEADER_ALIASES.keys())

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

SCHEMA = """
CREATE TABLE _meta (key TEXT PRIMARY KEY, value TEXT);

-- Org tables
CREATE TABLE org (
    org_id INTEGER PRIMARY KEY, org_key TEXT UNIQUE NOT NULL,
    org_name TEXT NOT NULL, tab_name TEXT NOT NULL
);
CREATE TABLE jira_component (
    component_id INTEGER PRIMARY KEY, component_name TEXT UNIQUE NOT NULL
);
CREATE TABLE scrum_team (
    team_id INTEGER PRIMARY KEY, team_name TEXT UNIQUE NOT NULL,
    pm TEXT, eng_lead TEXT
);
CREATE TABLE miro_team (
    miro_team_id INTEGER PRIMARY KEY, miro_team_name TEXT UNIQUE NOT NULL
);
CREATE TABLE specialty (
    specialty_id INTEGER PRIMARY KEY, specialty_name TEXT UNIQUE NOT NULL
);
CREATE TABLE person (
    person_id INTEGER PRIMARY KEY, name TEXT NOT NULL,
    manager TEXT, org_id INTEGER REFERENCES org(org_id),
    specialty_id INTEGER REFERENCES specialty(specialty_id),
    status TEXT, last_modified TEXT,
    user_id TEXT, job_title TEXT, email TEXT, location TEXT, manager_uid TEXT,
    source TEXT NOT NULL DEFAULT 'spreadsheet',
    UNIQUE(name, org_id)
);
CREATE TABLE person_component (
    person_id INTEGER REFERENCES person(person_id),
    component_id INTEGER REFERENCES jira_component(component_id),
    fte_fraction REAL NOT NULL,
    PRIMARY KEY (person_id, component_id)
);
CREATE TABLE person_miro_team (
    person_id INTEGER REFERENCES person(person_id),
    miro_team_id INTEGER REFERENCES miro_team(miro_team_id),
    PRIMARY KEY (person_id, miro_team_id)
);
CREATE TABLE person_scrum_team (
    person_id INTEGER REFERENCES person(person_id),
    team_id INTEGER REFERENCES scrum_team(team_id),
    PRIMARY KEY (person_id, team_id)
);
CREATE TABLE jira_scrum_mapping (
    component_name TEXT NOT NULL, scrum_team TEXT, specialty TEXT
);

-- Jira issues
CREATE TABLE jira_issue (
    issue_id INTEGER PRIMARY KEY, key TEXT UNIQUE NOT NULL,
    summary TEXT, status TEXT, priority TEXT,
    assignee TEXT, reporter TEXT, issue_type TEXT,
    created TEXT, updated TEXT
);
CREATE TABLE issue_label (
    issue_id INTEGER REFERENCES jira_issue(issue_id),
    label TEXT NOT NULL, PRIMARY KEY (issue_id, label)
);
CREATE TABLE issue_component (
    issue_id INTEGER REFERENCES jira_issue(issue_id),
    component_name TEXT NOT NULL, PRIMARY KEY (issue_id, component_name)
);

-- Feature planning / RICE scores
CREATE TABLE feature (
    feature_id INTEGER PRIMARY KEY, issue_key TEXT UNIQUE NOT NULL,
    title TEXT, project TEXT, hierarchy TEXT,
    assignee TEXT, sprint TEXT,
    target_start TEXT, target_end TEXT, due_date TEXT,
    estimates_days REAL, parent_key TEXT, priority TEXT,
    issue_status TEXT, progress_pct REAL,
    progress_completed_days REAL, progress_remaining_days REAL,
    progress_pct_ic REAL, todo_ic INTEGER, in_progress_ic INTEGER,
    done_ic INTEGER, total_ic INTEGER, color_status TEXT,
    release_date TEXT, rice_score REAL,
    dev_approval TEXT, developer TEXT, docs_approval TEXT,
    docs_impact TEXT, product_lead TEXT, product_manager TEXT,
    qe_approval TEXT, tester TEXT, owner TEXT, architect TEXT,
    plm_tech_lead TEXT, tech_lead TEXT, target_milestone TEXT
);
CREATE TABLE feature_release (
    feature_id INTEGER REFERENCES feature(feature_id),
    release TEXT NOT NULL, PRIMARY KEY (feature_id, release)
);
CREATE TABLE feature_label (
    feature_id INTEGER REFERENCES feature(feature_id),
    label TEXT NOT NULL, PRIMARY KEY (feature_id, label)
);
CREATE TABLE feature_component (
    feature_id INTEGER REFERENCES feature(feature_id),
    component TEXT NOT NULL, PRIMARY KEY (feature_id, component)
);
CREATE TABLE feature_team (
    feature_id INTEGER REFERENCES feature(feature_id),
    team TEXT NOT NULL, PRIMARY KEY (feature_id, team)
);

-- Release schedule
CREATE TABLE release_schedule (
    schedule_id INTEGER PRIMARY KEY,
    release TEXT NOT NULL, task TEXT NOT NULL,
    date_start TEXT, date_finish TEXT
);

-- Release-to-component version mapping
CREATE TABLE release_milestone (
    product TEXT NOT NULL, version TEXT NOT NULL,
    event_type TEXT NOT NULL, event_date TEXT,
    PRIMARY KEY (product, version, event_type)
);
CREATE TABLE component_version_map (
    component TEXT NOT NULL, release_version TEXT NOT NULL,
    component_version TEXT, component_group TEXT,
    PRIMARY KEY (component, release_version)
);

-- Org chart (raw hierarchical text)
CREATE TABLE org_chart_raw (
    line_num INTEGER PRIMARY KEY, content TEXT NOT NULL
);

-- Jira issue changelog (field change history)
CREATE TABLE issue_changelog (
    changelog_id INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_id INTEGER NOT NULL REFERENCES jira_issue(issue_id),
    field TEXT NOT NULL,
    field_type TEXT,
    from_value TEXT,
    to_value TEXT,
    author TEXT,
    changed_at TEXT NOT NULL
);

-- Convenience views
CREATE VIEW v_person_detail AS
SELECT
    p.person_id, p.name, p.user_id, p.manager, p.manager_uid,
    o.org_name, o.org_key, s.specialty_name AS specialty,
    p.job_title, p.email, p.location, p.status, p.source, p.last_modified,
    GROUP_CONCAT(DISTINCT jc.component_name) AS components,
    GROUP_CONCAT(DISTINCT mt.miro_team_name) AS miro_teams,
    GROUP_CONCAT(DISTINCT st.team_name) AS scrum_teams
FROM person p
LEFT JOIN org o ON p.org_id = o.org_id
LEFT JOIN specialty s ON p.specialty_id = s.specialty_id
LEFT JOIN person_component pc ON p.person_id = pc.person_id
LEFT JOIN jira_component jc ON pc.component_id = jc.component_id
LEFT JOIN person_miro_team pmt ON p.person_id = pmt.person_id
LEFT JOIN miro_team mt ON pmt.miro_team_id = mt.miro_team_id
LEFT JOIN person_scrum_team pst ON p.person_id = pst.person_id
LEFT JOIN scrum_team st ON pst.team_id = st.team_id
GROUP BY p.person_id;

CREATE VIEW v_component_headcount AS
SELECT jc.component_name, COUNT(pc.person_id) AS headcount,
       ROUND(SUM(pc.fte_fraction), 1) AS fte_total
FROM jira_component jc
LEFT JOIN person_component pc ON jc.component_id = pc.component_id
GROUP BY jc.component_id ORDER BY fte_total DESC;

CREATE VIEW v_team_headcount AS
SELECT st.team_name, COUNT(pst.person_id) AS headcount
FROM scrum_team st
LEFT JOIN person_scrum_team pst ON st.team_id = pst.team_id
GROUP BY st.team_id ORDER BY headcount DESC;

CREATE VIEW v_unassigned AS
SELECT p.person_id, p.name, p.user_id, p.job_title, p.email,
       p.location, p.manager_uid
FROM person p
WHERE p.person_id NOT IN (SELECT person_id FROM person_scrum_team)
  AND p.person_id NOT IN (SELECT person_id FROM person_component)
ORDER BY p.name;

CREATE VIEW v_issue_component_summary AS
SELECT ic.component_name, COUNT(*) AS issue_count,
       SUM(CASE WHEN ji.status IN ('Resolved','Closed','Done') THEN 1 ELSE 0 END) AS done_count
FROM jira_issue ji JOIN issue_component ic USING(issue_id)
GROUP BY ic.component_name ORDER BY issue_count DESC;

CREATE VIEW v_feature_summary AS
SELECT issue_status, COUNT(*) AS cnt, ROUND(AVG(rice_score),1) AS avg_rice
FROM feature GROUP BY issue_status ORDER BY cnt DESC;

-- Governance documents
CREATE TABLE governance_document (
    doc_id INTEGER PRIMARY KEY,
    doc_type TEXT NOT NULL CHECK(doc_type IN ('constitution','policy','standard','reference')),
    title TEXT NOT NULL,
    version TEXT,
    category TEXT,
    content TEXT,
    sections TEXT,  -- JSON: [{heading, level, content}]
    source_file TEXT,
    source_url TEXT,
    extracted_at TEXT,
    hash TEXT
);

CREATE VIEW v_governance_toc AS
SELECT doc_id, doc_type, title, version, category, source_file,
       json_group_array(json_object(
           'heading', json_extract(value, '$.heading'),
           'level', json_extract(value, '$.level')
       )) AS table_of_contents
FROM governance_document, json_each(sections)
GROUP BY doc_id;

-- Indexes: jira_issue (search_issues tool)
CREATE INDEX idx_jira_issue_status ON jira_issue(status);
CREATE INDEX idx_jira_issue_assignee ON jira_issue(assignee);
CREATE INDEX idx_jira_issue_priority ON jira_issue(priority);
CREATE INDEX idx_jira_issue_updated ON jira_issue(updated);
CREATE INDEX idx_issue_component_name ON issue_component(component_name);
CREATE INDEX idx_issue_label_label ON issue_label(label);
-- Indexes: person (lookup_person tool)
CREATE INDEX idx_person_name ON person(name);
CREATE INDEX idx_person_user_id ON person(user_id);
CREATE INDEX idx_person_email ON person(email);
-- Indexes: feature (get_feature_status, release_risk_summary tools)
CREATE INDEX idx_feature_issue_key ON feature(issue_key);
CREATE INDEX idx_feature_release_fid ON feature_release(feature_id);
CREATE INDEX idx_feature_component_fid ON feature_component(feature_id);
CREATE INDEX idx_feature_team_fid ON feature_team(feature_id);
-- Indexes: issue_changelog
CREATE INDEX idx_changelog_issue ON issue_changelog(issue_id);
CREATE INDEX idx_changelog_field ON issue_changelog(field);
CREATE INDEX idx_changelog_changed_at ON issue_changelog(changed_at);
-- Indexes: release + governance
CREATE INDEX idx_release_milestone_version ON release_milestone(version);
CREATE INDEX idx_governance_doc_type ON governance_document(doc_type);
"""

# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def split_multi_value(value: str | None) -> list[str]:
    if not value or not str(value).strip():
        return []
    return [v.strip() for v in str(value).split(",") if v.strip()]


def get_or_create(cur, table: str, name_col: str, name_val: str, id_col: str) -> int:
    cur.execute(f"INSERT OR IGNORE INTO {table} ({name_col}) VALUES (?)", (name_val,))
    cur.execute(f"SELECT {id_col} FROM {table} WHERE {name_col} = ?", (name_val,))
    return cur.fetchone()[0]


def normalize_name(name: str) -> str:
    name = re.sub(r"\s*\(.*?\)\s*", " ", name)
    return " ".join(name.lower().split())


def find_csv(directory: Path, pattern: str) -> Path | None:
    candidates = sorted(
        (f for f in directory.iterdir() if f.is_file() and f.suffix == ".csv" and pattern in f.name),
        key=lambda f: f.name,
        reverse=True,
    )
    return candidates[0] if candidates else None


def find_xlsx(directory: Path) -> Path | None:
    candidates = sorted(
        (f for f in directory.iterdir() if f.is_file() and f.suffix == ".xlsx"),
        key=lambda f: f.name,
    )
    if not candidates:
        return None
    # Prefer the file whose sheets match the configured TAB_ORG_MAP
    expected_tabs = set(TAB_ORG_MAP.keys())
    for c in candidates:
        try:
            wb = load_workbook(c, read_only=True)
            tabs = set(wb.sheetnames)
            wb.close()
            if expected_tabs & tabs:
                return c
        except Exception:
            continue
    return candidates[0]


# ---------------------------------------------------------------------------
# Org spreadsheet parsers
# ---------------------------------------------------------------------------


def _match_header(actual: str) -> str | None:
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if actual == alias or actual.startswith(alias):
                return canonical
    return None


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, 6):
        cells = {}
        for col_idx, cell in enumerate(ws[row_idx], start=0):
            val = str(cell.value or "").strip().lower()
            if val:
                matched = _match_header(val)
                if matched:
                    cells[matched] = col_idx
        if len(cells) >= 5:
            return row_idx, cells
    return 1, {}


def parse_team_breakdown(ws) -> list[dict]:
    header_row, col_map = find_header_row(ws)
    if not col_map:
        print(f"  WARNING: Could not find headers in '{ws.title}', skipping")
        return []
    people = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):

        def get(header):
            idx = col_map.get(header)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            if val is None:
                return None
            return str(val).strip() if not isinstance(val, datetime) else val.strftime("%Y-%m-%d")

        name = get("associate's name")
        if not name or name.lower() in ("", "associate's name"):
            continue
        people.append(
            {
                "name": name,
                "manager": get("manager's name"),
                "miro_team_name": get("miro team name"),
                "primary_jira_component": get("primary jira component"),
                "jira_team_name": get("jira team name"),
                "pm": get("pm"),
                "eng_lead": get("eng lead"),
                "status": get("status"),
                "specialty": get("engineering speciality"),
                "last_modified": get("last modified date"),
            }
        )
    return people


def parse_jira_scrum_ref(ws) -> list[dict]:
    rows = []
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    comp_idx = next((i for i, h in enumerate(headers) if "jira" in h and "component" in h), 0)
    scrum_idx = next((i for i, h in enumerate(headers) if "scrum" in h or "team" in h), 1)
    spec_idx = next((i for i, h in enumerate(headers) if "special" in h), 2)
    for row in ws.iter_rows(min_row=2, values_only=True):
        comp = str(row[comp_idx] or "").strip() if comp_idx < len(row) else ""
        scrum = str(row[scrum_idx] or "").strip() if scrum_idx < len(row) else ""
        spec = str(row[spec_idx] or "").strip() if spec_idx < len(row) else ""
        if comp:
            rows.append(
                {
                    "component_name": comp,
                    "scrum_team": scrum or None,
                    "specialty": spec or None,
                }
            )
    return rows


# ---------------------------------------------------------------------------
# CSV loaders
# ---------------------------------------------------------------------------


def load_jira_issues(cur, csv_path: Path) -> int:
    """Load Jira issues export CSV. Returns row count."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            key = row.get("key", "").strip()
            if not key:
                continue
            cur.execute(
                "INSERT OR IGNORE INTO jira_issue "
                "(key, summary, status, priority, assignee, reporter, issue_type, created, updated) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    key,
                    row.get("summary", ""),
                    row.get("status", ""),
                    row.get("priority", ""),
                    row.get("assignee", ""),
                    row.get("reporter", ""),
                    row.get("issuetype", ""),
                    row.get("created", ""),
                    row.get("updated", ""),
                ),
            )
            cur.execute("SELECT issue_id FROM jira_issue WHERE key = ?", (key,))
            issue_id = cur.fetchone()[0]
            for label in split_multi_value(row.get("labels")):
                cur.execute(
                    "INSERT OR IGNORE INTO issue_label (issue_id, label) VALUES (?, ?)",
                    (issue_id, label),
                )
            for comp in split_multi_value(row.get("components")):
                cur.execute(
                    "INSERT OR IGNORE INTO issue_component (issue_id, component_name) VALUES (?, ?)",
                    (issue_id, comp),
                )
            count += 1
    return count


def load_jira_changelog(cur, csv_path: Path) -> int:
    """Load Jira changelog CSV. Returns row count."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            key = row.get("key", "").strip()
            if not key:
                continue
            cur.execute("SELECT issue_id FROM jira_issue WHERE key = ?", (key,))
            result = cur.fetchone()
            if not result:
                continue
            issue_id = result[0]
            cur.execute(
                "INSERT INTO issue_changelog (issue_id, field, field_type, from_value, to_value, author, changed_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    issue_id,
                    row.get("field", ""),
                    row.get("field_type", ""),
                    row.get("from_value", ""),
                    row.get("to_value", ""),
                    row.get("author", ""),
                    row.get("changed_at", ""),
                ),
            )
            count += 1
    return count


def _parse_float(val: str) -> float | None:
    try:
        return float(val) if val and val.strip() else None
    except ValueError:
        return None


def _parse_int(val: str) -> int | None:
    try:
        return int(float(val)) if val and val.strip() else None
    except ValueError:
        return None


def load_features(cur, csv_path: Path) -> int:
    """Load feature planning / RICE scores CSV. Returns row count."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            issue_key = row.get("Issue key", "").strip()
            if not issue_key:
                continue
            cur.execute(
                "INSERT OR IGNORE INTO feature (issue_key, title, project, hierarchy, assignee, sprint, "
                "target_start, target_end, due_date, estimates_days, parent_key, priority, issue_status, "
                "progress_pct, progress_completed_days, progress_remaining_days, progress_pct_ic, "
                "todo_ic, in_progress_ic, done_ic, total_ic, color_status, release_date, rice_score, "
                "dev_approval, developer, docs_approval, docs_impact, product_lead, product_manager, "
                "qe_approval, tester, owner, architect, plm_tech_lead, tech_lead, target_milestone) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    issue_key,
                    row.get("Title", ""),
                    row.get("Project", ""),
                    row.get("Hierarchy", ""),
                    row.get("Assignee", ""),
                    row.get("Sprint", ""),
                    row.get("Target start date", ""),
                    row.get("Target end date", ""),
                    row.get("Due date", ""),
                    _parse_float(row.get("Estimates (d)", "")),
                    row.get("Parent", ""),
                    row.get("Priority", ""),
                    row.get("Issue status", ""),
                    _parse_float(row.get("Progress (%)", "")),
                    _parse_float(row.get("Progress completed (d)", "")),
                    _parse_float(row.get("Progress remaining (d)", "")),
                    _parse_float(row.get("Progress (%) issue count (IC)", "")),
                    _parse_int(row.get("To do IC", "")),
                    _parse_int(row.get("In progress IC", "")),
                    _parse_int(row.get("Done IC", "")),
                    _parse_int(row.get("Total IC", "")),
                    row.get("Color Status", ""),
                    row.get("Release Date", ""),
                    _parse_float(row.get("RICE Score", "")),
                    row.get("Dev Approval", ""),
                    row.get("Developer", ""),
                    row.get("Docs Approval", ""),
                    row.get("Docs Impact", ""),
                    row.get("Product Lead", ""),
                    row.get("Product Manager", ""),
                    row.get("QE Approval", ""),
                    row.get("Tester", ""),
                    row.get("Owner", ""),
                    row.get("Architect", ""),
                    row.get("PLM  Technical Lead", ""),
                    row.get("Technical Lead", ""),
                    row.get("Target Milestone", ""),
                ),
            )
            cur.execute("SELECT feature_id FROM feature WHERE issue_key = ?", (issue_key,))
            feature_id = cur.fetchone()[0]
            for rel in split_multi_value(row.get("Releases")):
                cur.execute(
                    "INSERT OR IGNORE INTO feature_release (feature_id, release) VALUES (?, ?)",
                    (feature_id, rel),
                )
            for label in split_multi_value(row.get("Labels")):
                cur.execute(
                    "INSERT OR IGNORE INTO feature_label (feature_id, label) VALUES (?, ?)",
                    (feature_id, label),
                )
            for comp in split_multi_value(row.get("Components")):
                cur.execute(
                    "INSERT OR IGNORE INTO feature_component (feature_id, component) VALUES (?, ?)",
                    (feature_id, comp),
                )
            for team in split_multi_value(row.get("Team")):
                cur.execute(
                    "INSERT OR IGNORE INTO feature_team (feature_id, team) VALUES (?, ?)",
                    (feature_id, team),
                )
            count += 1
    return count


def load_release_schedule(cur, csv_path: Path) -> int:
    """Load release schedule CSV. Returns row count."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            release = row.get("Release", "").strip()
            task = row.get("Task", "").strip()
            if not release or not task:
                continue
            cur.execute(
                "INSERT INTO release_schedule (release, task, date_start, date_finish) VALUES (?, ?, ?, ?)",
                (release, task, row.get("Date Start", ""), row.get("Date Finish", "")),
            )
            count += 1
    return count


def load_release_version_map(cur, csv_path: Path) -> tuple[int, int]:
    """Load release-to-component version mapping CSV (pivot format).

    Returns (milestone_count, component_version_count).
    """
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    if not rows:
        return 0, 0

    # Row 0: product name + version headers
    versions = [v.strip() for v in rows[0][1:] if v.strip()]

    milestones = 0
    comp_versions = 0
    current_product = None
    current_group = None
    in_components = False

    for row in rows:
        if not row or all(not c.strip() for c in row):
            continue

        col0 = row[0].strip() if row else ""
        # Last non-empty cell might be an event type
        last_val = ""
        for c in reversed(row):
            if c.strip():
                last_val = c.strip()
                break

        # Detect product name rows (col0 has a name, no data in other columns)
        if (
            col0
            and all(not (row[i].strip() if i < len(row) else "") for i in range(1, min(4, len(row))))
            and col0 not in ("Components", "Dependencies")
        ):
            if not any(kw in col0.lower() for kw in ("components", "dependencies")):
                current_product = col0
                in_components = False
            continue

        # Product milestone rows (empty col0 or product name, event type in last col)
        if last_val in ("Major Version due", "Code freeze", "Release"):
            if col0 and not in_components:
                current_product = col0
                in_components = False
            if current_product:
                for i, ver in enumerate(versions):
                    val = row[i + 1].strip() if i + 1 < len(row) else ""
                    if val and val != "NA":
                        cur.execute(
                            "INSERT OR IGNORE INTO release_milestone "
                            "(product, version, event_type, event_date) VALUES (?, ?, ?, ?)",
                            (current_product, ver, last_val, val),
                        )
                        milestones += 1
            continue

        # (Continuation milestone rows are handled above)

        # Section headers
        if col0 and all(not (row[i].strip() if i < len(row) else "") for i in range(1, min(4, len(row)))):
            # Any section header triggers component mode
            current_group = col0
            in_components = True
            continue

        # Component version rows
        if in_components and col0:
            for i, ver in enumerate(versions):
                val = row[i + 1].strip() if i + 1 < len(row) else ""
                if val:
                    cur.execute(
                        "INSERT OR IGNORE INTO component_version_map "
                        "(component, release_version, component_version, component_group) "
                        "VALUES (?, ?, ?, ?)",
                        (col0, ver, val, current_group),
                    )
                    comp_versions += 1

    return milestones, comp_versions


def load_org_chart_raw(cur, csv_path: Path) -> int:
    """Load org chart as raw lines. Returns line count."""
    with open(csv_path, encoding="utf-8") as f:
        lines = f.readlines()
    count = 0
    for i, line in enumerate(lines, start=1):
        content = line.strip().strip('"')
        if content:
            cur.execute(
                "INSERT INTO org_chart_raw (line_num, content) VALUES (?, ?)",
                (i, content),
            )
            count += 1
    return count


# ---------------------------------------------------------------------------
# Governance document loaders
# ---------------------------------------------------------------------------


def parse_markdown_sections(text: str) -> list[dict]:
    """Split markdown text by headings into [{heading, level, content}]."""
    sections = []
    current = None
    for line in text.splitlines():
        match = re.match(r"^(#{1,6})\s+(.+)$", line)
        if match:
            if current:
                current["content"] = current["content"].strip()
                sections.append(current)
            current = {
                "heading": match.group(2).strip(),
                "level": len(match.group(1)),
                "content": "",
            }
        elif current:
            current["content"] += line + "\n"
        else:
            # Content before first heading
            current = {"heading": "(preamble)", "level": 0, "content": line + "\n"}
    if current:
        current["content"] = current["content"].strip()
        sections.append(current)
    return sections


def load_governance_policies(cur, governance_dir: Path) -> None:
    """Load governance/*.md as policy documents."""
    if not governance_dir.exists():
        return
    loaded = 0
    for md_path in sorted(governance_dir.glob("*.md")):
        text = md_path.read_text()
        file_hash = file_sha256(md_path)

        # Check if already loaded with same hash
        cur.execute(
            "SELECT doc_id FROM governance_document WHERE source_file = ? AND hash = ?",
            (md_path.name, file_hash),
        )
        if cur.fetchone():
            continue

        # Extract title from first heading
        title_match = re.match(r"^#\s+(.+)$", text, re.MULTILINE)
        title = title_match.group(1).strip() if title_match else md_path.stem

        sections = parse_markdown_sections(text)
        cur.execute(
            "INSERT INTO governance_document "
            "(doc_type, title, content, sections, source_file, extracted_at, hash) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                "policy",
                title,
                text,
                json.dumps(sections),
                md_path.name,
                datetime.now().isoformat(),
                file_hash,
            ),
        )
        loaded += 1
    if loaded:
        print(f"\nGovernance policies: {loaded} loaded")


def load_governance_pdfs(cur, data_dir: Path) -> None:
    """Load PDFs from data/ via docling, extracting text and sections."""
    pdf_files = sorted(data_dir.glob("*.pdf"))
    if not pdf_files:
        return

    # Check which PDFs need processing
    to_process = []
    for pdf_path in pdf_files:
        pdf_hash = file_sha256(pdf_path)
        cur.execute(
            "SELECT doc_id FROM governance_document WHERE source_file = ? AND hash = ?",
            (pdf_path.name, pdf_hash),
        )
        if not cur.fetchone():
            to_process.append((pdf_path, pdf_hash))

    if not to_process:
        print(f"\nPDFs: {len(pdf_files)} already up to date")
        return

    print(f"\nPDFs: extracting {len(to_process)} of {len(pdf_files)} via docling...")

    try:
        from concurrent.futures import ProcessPoolExecutor

        with ProcessPoolExecutor() as executor:
            futures = {executor.submit(_extract_pdf, p): (p, h) for p, h in to_process}
            for future in futures:
                pdf_path, pdf_hash = futures[future]
                try:
                    markdown_text = future.result(timeout=300)
                    if not markdown_text:
                        print(f"  WARNING: empty extraction for {pdf_path.name}")
                        continue
                    sections = parse_markdown_sections(markdown_text)
                    title_match = re.match(r"^#\s+(.+)$", markdown_text, re.MULTILINE)
                    title = title_match.group(1).strip() if title_match else pdf_path.stem

                    cur.execute(
                        "INSERT INTO governance_document "
                        "(doc_type, title, content, sections, source_file, extracted_at, hash) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (
                            "reference",
                            title,
                            markdown_text,
                            json.dumps(sections),
                            pdf_path.name,
                            datetime.now().isoformat(),
                            pdf_hash,
                        ),
                    )
                    print(f"  Loaded: {pdf_path.name}")
                except Exception as exc:
                    print(f"  ERROR extracting {pdf_path.name}: {exc}")
    except ImportError:
        print("  WARNING: docling not available, skipping PDF extraction")


def _extract_pdf(pdf_path: Path) -> str:
    """Worker function: convert a single PDF to markdown via docling."""
    from docling.document_converter import DocumentConverter

    converter = DocumentConverter()
    result = converter.convert(str(pdf_path))
    return result.document.export_to_markdown()


def check_catalog_sync(data_dir: Path, catalog_path: Path) -> None:
    """Warn if DATA_CATALOG.yaml entries don't match actual files in data/."""
    if not catalog_path.exists():
        return
    catalog_text = catalog_path.read_text()
    catalog_files = set(re.findall(r'(?:filename|file):\s*"?([^"\n]+)"?', catalog_text))
    disk_files = {f.name for f in data_dir.iterdir() if f.is_file() and f.suffix in (".csv", ".xlsx", ".pdf")}

    missing_from_catalog = disk_files - catalog_files
    missing_from_disk = catalog_files - disk_files

    if missing_from_catalog:
        print(f"\nCatalog sync: {len(missing_from_catalog)} file(s) on disk not in catalog:")
        for f in sorted(missing_from_catalog):
            print(f"  + {f}")
    if missing_from_disk:
        print(f"\nCatalog sync: {len(missing_from_disk)} catalog entries with no file:")
        for f in sorted(missing_from_disk):
            print(f"  ? {f}")


# ---------------------------------------------------------------------------
# Main build orchestrator
# ---------------------------------------------------------------------------


def build_database(xlsx_path: Path, db_path: Path) -> None:
    print(f"Loading workbook: {xlsx_path.name}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  Found {len(sheet_names)} tabs: {', '.join(sheet_names[:8])}{'...' if len(sheet_names) > 8 else ''}")

    if db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = OFF")
    conn.execute("PRAGMA cache_size = -64000")
    conn.execute("PRAGMA temp_store = MEMORY")
    cur = conn.cursor()
    cur.executescript(SCHEMA)

    source_hash = file_sha256(xlsx_path)
    cur.execute("INSERT INTO _meta VALUES ('source_file', ?)", (xlsx_path.name,))
    cur.execute("INSERT INTO _meta VALUES ('source_hash', ?)", (source_hash,))
    cur.execute("INSERT INTO _meta VALUES ('built_at', ?)", (datetime.now().isoformat(),))

    # --- Org data ---
    for tab_name, (org_key, org_name) in TAB_ORG_MAP.items():
        cur.execute(
            "INSERT INTO org (org_key, org_name, tab_name) VALUES (?, ?, ?)",
            (org_key, org_name, tab_name),
        )

    total_people = 0
    total_multi_component = 0
    for tab_name, (org_key, org_name) in TAB_ORG_MAP.items():
        if tab_name not in sheet_names:
            print(f"  WARNING: Tab '{tab_name}' not found, skipping")
            continue
        ws = wb[tab_name]
        people = parse_team_breakdown(ws)
        print(f"  {tab_name}: {len(people)} people")
        cur.execute("SELECT org_id FROM org WHERE org_key = ?", (org_key,))
        org_id = cur.fetchone()[0]

        for person in people:
            specialty_id = None
            if person["specialty"]:
                specialty_id = get_or_create(
                    cur,
                    "specialty",
                    "specialty_name",
                    person["specialty"],
                    "specialty_id",
                )
            cur.execute(
                "INSERT OR IGNORE INTO person "
                "(name, manager, org_id, specialty_id, status, last_modified, source) "
                "VALUES (?, ?, ?, ?, ?, ?, 'spreadsheet')",
                (
                    person["name"],
                    person["manager"],
                    org_id,
                    specialty_id,
                    person["status"],
                    person["last_modified"],
                ),
            )
            cur.execute(
                "SELECT person_id FROM person WHERE name = ? AND org_id = ?",
                (person["name"], org_id),
            )
            person_id = cur.fetchone()[0]
            total_people += 1

            components = split_multi_value(person["primary_jira_component"])
            components = list(dict.fromkeys(components))
            fte = 1.0 / len(components) if components else 1.0
            if len(components) > 1:
                total_multi_component += 1
            for comp_name in components:
                comp_id = get_or_create(cur, "jira_component", "component_name", comp_name, "component_id")
                cur.execute(
                    "INSERT OR IGNORE INTO person_component (person_id, component_id, fte_fraction) VALUES (?, ?, ?)",
                    (person_id, comp_id, fte),
                )

            miro_teams = split_multi_value(person["miro_team_name"])
            miro_teams = list(dict.fromkeys(miro_teams))
            for mt_name in miro_teams:
                mt_id = get_or_create(cur, "miro_team", "miro_team_name", mt_name, "miro_team_id")
                cur.execute(
                    "INSERT OR IGNORE INTO person_miro_team (person_id, miro_team_id) VALUES (?, ?)",
                    (person_id, mt_id),
                )

            scrum_teams = split_multi_value(person["jira_team_name"])
            scrum_teams = list(dict.fromkeys(scrum_teams))
            for st_name in scrum_teams:
                cur.execute(
                    "INSERT OR IGNORE INTO scrum_team (team_name, pm, eng_lead) VALUES (?, ?, ?)",
                    (st_name, person["pm"], person["eng_lead"]),
                )
                cur.execute("SELECT team_id FROM scrum_team WHERE team_name = ?", (st_name,))
                team_id = cur.fetchone()[0]
                cur.execute(
                    "INSERT OR IGNORE INTO person_scrum_team (person_id, team_id) VALUES (?, ?)",
                    (person_id, team_id),
                )

    if JIRA_SCRUM_REF_TAB in sheet_names:
        ws = wb[JIRA_SCRUM_REF_TAB]
        ref_rows = parse_jira_scrum_ref(ws)
        print(f"  {JIRA_SCRUM_REF_TAB}: {len(ref_rows)} mappings")
        for row in ref_rows:
            cur.execute(
                "INSERT INTO jira_scrum_mapping (component_name, scrum_team, specialty) VALUES (?, ?, ?)",
                (row["component_name"], row["scrum_team"], row["specialty"]),
            )
    wb.close()

    # --- CSV data sources ---
    csv_sources = [
        ("acme-issues", "Jira issues", load_jira_issues),
        ("acme-feature", "Feature planning", load_features),
        ("acme-release-schedule", "Release schedule", load_release_schedule),
        ("acme-changelog", "Jira changelog", load_jira_changelog),
    ]

    for pattern, label, loader in csv_sources:
        csv_path = find_csv(DATA_DIR, pattern)
        if csv_path:
            result = loader(cur, csv_path)
            print(f"\n{label}: {csv_path.name}")
            print(f"  {result} rows loaded")
            cur.execute(
                "INSERT INTO _meta VALUES (?, ?)",
                (f"csv_{pattern}_hash", file_sha256(csv_path)),
            )
        else:
            print(f"\n{label}: not found (pattern: *{pattern}*)")

    # Release-to-component version map (special return type)
    ver_csv = find_csv(DATA_DIR, "acme-component-versions")
    if ver_csv:
        milestones, comp_vers = load_release_version_map(cur, ver_csv)
        print(f"\nRelease version map: {ver_csv.name}")
        print(f"  {milestones} milestones, {comp_vers} component versions")
        cur.execute(
            "INSERT INTO _meta VALUES (?, ?)",
            ("csv_component-versions_hash", file_sha256(ver_csv)),
        )
    else:
        print("\nRelease version map: not found")

    # --- GPS version ---
    if VERSION_PATH.exists():
        gps_version = VERSION_PATH.read_text().strip()
        cur.execute("INSERT OR REPLACE INTO _meta VALUES ('gps_version', ?)", (gps_version,))
        print(f"\nGPS version: {gps_version}")

    # --- Governance documents ---
    load_governance_policies(cur, GOVERNANCE_DIR)
    load_governance_pdfs(cur, DATA_DIR)

    # --- Catalog sync check ---
    catalog_path = DATA_DIR / "DATA_CATALOG.yaml"
    check_catalog_sync(DATA_DIR, catalog_path)

    conn.commit()

    # --- Post-build integrity ---
    result = conn.execute("PRAGMA integrity_check").fetchone()
    if result[0] != "ok":
        print(f"\nWARNING: integrity check failed: {result[0]}")
    conn.execute("ANALYZE")
    conn.close()

    # --- Summary ---
    conn = sqlite3.connect(str(db_path))
    counts = {}
    for tbl in (
        "person",
        "jira_issue",
        "feature",
        "release_schedule",
        "release_milestone",
        "component_version_map",
        "org_chart_raw",
        "governance_document",
        "issue_changelog",
    ):
        counts[tbl] = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
    conn.close()

    db_size = db_path.stat().st_size
    print(f"\nDatabase built: {db_path.name} ({db_size:,} bytes)")
    for tbl, cnt in counts.items():
        print(f"  {tbl}: {cnt:,}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build unified SQLite DB from organizational data sources.")
    parser.add_argument("--xlsx", type=Path, help="Path to XLSX file (auto-detected if omitted)")
    parser.add_argument(
        "--db",
        type=Path,
        default=DATA_DIR / "gps.db",
        help="Output database path (default: data/gps.db)",
    )
    parser.add_argument("--force", action="store_true", help="Rebuild even if sources haven't changed")
    args = parser.parse_args()

    xlsx_path = args.xlsx or find_xlsx(DATA_DIR)
    if not xlsx_path or not xlsx_path.exists():
        print(
            "ERROR: No XLSX file found in data/. Place an .xlsx org spreadsheet there first.",
            file=sys.stderr,
        )
        sys.exit(1)

    # Hash-based skip
    if args.db.exists() and not args.force:
        try:
            conn = sqlite3.connect(str(args.db))
            cur = conn.cursor()
            cur.execute("SELECT value FROM _meta WHERE key = 'source_hash'")
            row = cur.fetchone()
            xlsx_unchanged = row and row[0] == file_sha256(xlsx_path)
            conn.close()
            if xlsx_unchanged:
                print("Sources unchanged (hashes match), skipping rebuild. Use --force to override.")
                sys.exit(0)
        except Exception:
            pass

    build_database(xlsx_path, args.db)


if __name__ == "__main__":
    main()
